import pandas as pd
import numpy as np
from datetime import timedelta
import calendar
import warnings
warnings.filterwarnings("ignore")
PREDICT_DAYS = 3      
MONTH_WINDOW = 1      
START_DATE    = "2026-04-01" 
N_TRIALS     = 300      
PERIODS = {
    "Night"     : (0,  5),
    "Morning"   : (6,  11),
    "Afternoon" : (12, 17),
    "Evening"   : (18, 23),
}
try:
    from xgboost import XGBRegressor
except ImportError:
    raise ImportError("pip install xgboost")

try:
    import optuna
    optuna.logging.set_verbosity(optuna.logging.WARNING)
except ImportError:
    raise ImportError("pip install optuna")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
except ImportError:
    raise ImportError("pip install openpyxl")

from sklearn.model_selection import TimeSeriesSplit
from sklearn.metrics import mean_squared_error
print("=" * 60)
print(" Power Prediction — XGBoost + Optuna")
print("=" * 60)
print("\n[1/6] Loading Dataset.xlsx ...")
df = pd.read_excel("Dataset.xlsx", parse_dates=["DateTime"])
df = df.sort_values("DateTime").reset_index(drop=True)
df = df.dropna(subset=["Load_MW"])
print(f"  Records : {len(df):,}")
print(f"  Range   : {df['DateTime'].min().date()} → {df['DateTime'].max().date()}")
print(f"  Load    : {df['Load_MW'].min():.1f} – {df['Load_MW'].max():.1f} MW")
def is_weekend(ts):
    return ts.weekday() >= 5
def get_period_name(h):
    for name, (s, e) in PERIODS.items():
        if s <= h <= e:
            return name
    return "Night"
print("\n[2/6] Engineering features ...")
df["hour"]       = df["DateTime"].dt.hour
df["minute"]     = df["DateTime"].dt.minute
df["month"]      = df["DateTime"].dt.month
df["slot"]       = df["hour"] * 4 + df["minute"] // 15
df["dayofweek"]  = df["DateTime"].dt.dayofweek
df["is_weekend"] = df["DateTime"].apply(is_weekend).astype(int)
df["period"]     = df["hour"].apply(get_period_name)

slot_stats = (
    df.groupby("slot")["Load_MW"]
    .agg(slot_mean="mean", slot_std="std")
    .reset_index()
    .fillna(0)
)
def build_features(d, slot_stats):
    d  = d.copy()
    dt = d["DateTime"]

    d["hour"]        = dt.dt.hour
    d["minute"]      = dt.dt.minute
    d["dayofweek"]   = dt.dt.dayofweek
    d["is_weekend"]  = d["DateTime"].apply(is_weekend).astype(int)
    d["month"]       = dt.dt.month
    d["quarter"]     = dt.dt.quarter
    d["dayofyear"]   = dt.dt.dayofyear
    d["weekofyear"]  = dt.dt.isocalendar().week.astype(int)
    d["slot"]        = d["hour"] * 4 + d["minute"] // 15

    d["hour_sin"]    = np.sin(2 * np.pi * d["hour"]      / 24)
    d["hour_cos"]    = np.cos(2 * np.pi * d["hour"]      / 24)
    d["slot_sin"]    = np.sin(2 * np.pi * d["slot"]      / 96)
    d["slot_cos"]    = np.cos(2 * np.pi * d["slot"]      / 96)
    d["dow_sin"]     = np.sin(2 * np.pi * d["dayofweek"] / 7)
    d["dow_cos"]     = np.cos(2 * np.pi * d["dayofweek"] / 7)
    d["month_sin"]   = np.sin(2 * np.pi * d["month"]     / 12)
    d["month_cos"]   = np.cos(2 * np.pi * d["month"]     / 12)

    for lag_days in [1, 2, 3, 7, 14]:
        d[f"lag_{lag_days}d"] = d["Load_MW"].shift(lag_days * 96)

    d["roll_mean_1d"] = d["Load_MW"].shift(1).rolling(96,  min_periods=1).mean()
    d["roll_std_1d"]  = d["Load_MW"].shift(1).rolling(96,  min_periods=1).std().fillna(0)
    d["roll_mean_7d"] = d["Load_MW"].shift(1).rolling(672, min_periods=1).mean()

    d = d.merge(slot_stats, on="slot", how="left")
    return d

df = build_features(df, slot_stats)

FEATURE_COLS = [
    "hour", "minute", "dayofweek", "is_weekend",
    "month", "quarter", "dayofyear", "weekofyear", "slot",
    "hour_sin", "hour_cos", "slot_sin", "slot_cos",
    "dow_sin",  "dow_cos",  "month_sin", "month_cos",
    "lag_1d", "lag_2d", "lag_3d", "lag_7d", "lag_14d",
    "roll_mean_1d", "roll_std_1d", "roll_mean_7d",
    "slot_mean", "slot_std",
]
def get_context_data(df, target_month, period_name, day_type, window=MONTH_WINDOW):
    h_start, h_end = PERIODS[period_name]
    MIN_ROWS = 200

    mask = (
        df["month"].isin({((target_month - 1 + o) % 12) + 1
                          for o in range(-window, window + 1)})
        & (df["is_weekend"] == (1 if day_type == "weekend" else 0))
        & (df["hour"] >= h_start) & (df["hour"] <= h_end)
    )
    ctx = df[mask].dropna(subset=FEATURE_COLS + ["Load_MW"])

    if len(ctx) < MIN_ROWS:
        mask2 = (
            df["month"].isin({((target_month - 1 + o) % 12) + 1
                              for o in range(-window, window + 1)})
            & (df["hour"] >= h_start) & (df["hour"] <= h_end)
        )
        ctx = df[mask2].dropna(subset=FEATURE_COLS + ["Load_MW"])

    if len(ctx) < MIN_ROWS:
        mask3 = (df["hour"] >= h_start) & (df["hour"] <= h_end)
        ctx = df[mask3].dropna(subset=FEATURE_COLS + ["Load_MW"])

    return ctx
def tune_and_train(ctx_df, label="model"):
    X = ctx_df[FEATURE_COLS].values
    y = ctx_df["Load_MW"].values
    tscv = TimeSeriesSplit(n_splits=3)

    trial_history = []  

    def objective(trial):
        params = {
            "n_estimators"    : trial.suggest_int  ("n_estimators",     200, 1500),
            "learning_rate"   : trial.suggest_float("learning_rate",    0.01, 0.3,  log=True),
            "max_depth"       : trial.suggest_int  ("max_depth",        3, 10),
            "min_child_weight": trial.suggest_int  ("min_child_weight", 1, 10),
            "subsample"       : trial.suggest_float("subsample",        0.5, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.4, 1.0),
            "reg_alpha"       : trial.suggest_float("reg_alpha",        1e-4, 10.0, log=True),
            "reg_lambda"      : trial.suggest_float("reg_lambda",       1e-4, 10.0, log=True),
            "gamma"           : trial.suggest_float("gamma",            0.0, 5.0),
            "random_state"    : 42,
            "n_jobs"          : -1,
            "verbosity"       : 0,
        }

        fold_rmses = []
        for train_idx, val_idx in tscv.split(X):
            mdl = XGBRegressor(**params)
            mdl.fit(X[train_idx], y[train_idx],
                    eval_set=[(X[val_idx], y[val_idx])],
                    verbose=False)
            preds = mdl.predict(X[val_idx])
            fold_rmses.append(np.sqrt(mean_squared_error(y[val_idx], preds)))

        cv_rmse = float(np.mean(fold_rmses))
        trial_history.append({
            "trial"           : trial.number + 1,
            "cv_rmse"         : round(cv_rmse, 3),
            **{k: round(v, 6) if isinstance(v, float) else v
               for k, v in params.items()
               if k not in ("random_state", "n_jobs", "verbosity")}
        })
        return cv_rmse

    study = optuna.create_study(
        direction="minimize",
        sampler=optuna.samplers.TPESampler(seed=42),
        pruner=optuna.pruners.MedianPruner(n_startup_trials=10, n_warmup_steps=5),
    )
    study.optimize(objective, n_trials=N_TRIALS, show_progress_bar=False)

    best_params = study.best_params
    best_cv_rmse = round(study.best_value, 3)
    best_params.update({"random_state": 42, "n_jobs": -1, "verbosity": 0})
    split = ctx_df["DateTime"].max() - timedelta(days=7)
    train = ctx_df[ctx_df["DateTime"] <= split]
    val   = ctx_df[ctx_df["DateTime"] >  split]
    if len(val) < 10:
        train = ctx_df
        val   = ctx_df.tail(max(10, len(ctx_df) // 10))

    final_model = XGBRegressor(**best_params)
    final_model.fit(
        train[FEATURE_COLS], train["Load_MW"],
        eval_set=[(val[FEATURE_COLS], val["Load_MW"])],
        verbose=False,
    )

    val_preds = final_model.predict(val[FEATURE_COLS])
    val_mae   = round(float(np.mean(np.abs(val_preds - val["Load_MW"]))), 3)
    val_mape  = round(float(np.mean(
        np.abs((val_preds - val["Load_MW"]) / val["Load_MW"])) * 100), 3)

    return final_model, best_params, best_cv_rmse, val_mae, val_mape, trial_history
last_date  = df["DateTime"].max().normalize()
start_date = pd.Timestamp(START_DATE) if START_DATE else last_date + timedelta(days=1)
total_slots  = PREDICT_DAYS * 96
future_times = pd.date_range(start_date, periods=total_slots, freq="15min")

needed_combos = {
    (ts.month, get_period_name(ts.hour), "weekend" if is_weekend(ts) else "weekday")
    for ts in future_times
}

print(f"\n[3/6] Tuning {len(needed_combos)} context model(s)  "
      f"({N_TRIALS} Optuna trials each) ...")
print(f"      Predicting: {start_date.date()} → "
      f"{(start_date + timedelta(days=PREDICT_DAYS-1)).date()}\n")

model_cache   = {}
model_meta    = {}
all_trials    = {}

for combo_idx, (month, period, day_type) in enumerate(sorted(needed_combos), 1):
    key   = (month, period, day_type)
    label = (f"  [{combo_idx}/{len(needed_combos)}]  "
             f"Month={calendar.month_abbr[month]}  "
             f"Period={period:<12s}  Type={day_type:<8s}")
    print(label, end="", flush=True)

    ctx = get_context_data(df, month, period, day_type)
    print(f"  {len(ctx):,} rows  → tuning ...", end="", flush=True)

    mdl, best_p, cv_rmse, val_mae, val_mape, trials = tune_and_train(ctx, label)

    model_cache[key] = mdl
    model_meta[key]  = {
        "rows"    : len(ctx),
        "cv_rmse" : cv_rmse,
        "val_mae" : val_mae,
        "val_mape": val_mape,
        "params"  : best_p,
    }
    all_trials[key] = trials

    print(f"  ✓  CV-RMSE={cv_rmse} MW  MAE={val_mae} MW  MAPE={val_mape}%")
print(f"\n[4/6] Generating {PREDICT_DAYS}-day predictions ...")

future_df = pd.DataFrame({"DateTime": future_times, "Load_MW": np.nan})
working   = pd.concat([df[["DateTime","Load_MW"]], future_df], ignore_index=True).copy()
fut_idx   = working[working["DateTime"] == future_times[0]].index[0]

for step, ts in enumerate(future_times):
    idx      = fut_idx + step
    period   = get_period_name(ts.hour)
    day_type = "weekend" if is_weekend(ts) else "weekday"
    mdl      = model_cache[(ts.month, period, day_type)]

    temp = build_features(working.iloc[:idx + 1].copy(), slot_stats)
    row  = temp.iloc[[-1]][FEATURE_COLS].copy()
    for col in FEATURE_COLS:
        if row[col].isnull().any():
            row[col] = df[col].median() if col in df.columns else 0

    pred = float(np.clip(mdl.predict(row)[0], 0, None))
    working.at[idx, "Load_MW"] = pred

    if (step + 1) % 96 == 0:
        d_num = (step + 1) // 96
        dp    = working.iloc[fut_idx + (d_num-1)*96 : fut_idx + d_num*96]["Load_MW"]
        dd    = start_date + timedelta(days=d_num - 1)
        print(f"  Day {d_num} ({dd.date()}) "
              f"[{'Weekend' if is_weekend(dd) else 'Weekday'}]  —  "
              f"avg={dp.mean():.1f} MW  peak={dp.max():.1f} MW  "
              f"energy={dp.sum()*0.25:.1f} MWh")

all_predictions = working.iloc[fut_idx : fut_idx + total_slots]["Load_MW"].values
end_date    = start_date + timedelta(days=PREDICT_DAYS - 1)
output_file = f"Power_Prediction_{start_date.date()}_to_{end_date.date()}.xlsx"
print(f"\n[5/6] Writing → {output_file} ...")

C_HDR = "1F4E79";  C_FG = "FFFFFF";  C_SUB = "2E75B6"
DAY_COLORS = ["2E75B6","70AD47","ED7D31","7030A0","C00000","00B0F0","FF6699"]
PERIOD_COLORS = {
    "Night"    : "D9E1F2", "Morning"  : "E2EFDA",
    "Afternoon": "FFF2CC", "Evening"  : "FCE4D6",
}
WEEKEND_COLOR = "F4CCFF"

thin = Side(style="thin", color="B8CCE4")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, bg=C_HDR, fg=C_FG, sz=11):
    cell.font      = Font(bold=True, color=fg, size=sz, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = bdr

def dat(cell, bg="FFFFFF"):
    cell.font      = Font(size=10, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = bdr

def row_bg(ts):
    if is_weekend(ts):
        return WEEKEND_COLOR
    return PERIOD_COLORS.get(get_period_name(ts.hour), "FFFFFF")

wb = Workbook()
ws1 = wb.active;  ws1.title = "All Predictions"
ws1.merge_cells("A1:G1")
tc = ws1["A1"]
tc.value     = (f"Context-Aware Power Prediction  ·  "
                f"{start_date.strftime('%d %b %Y')} → {end_date.strftime('%d %b %Y')}"
                f"  ({PREDICT_DAYS} day(s), Optuna-tuned)")
tc.font      = Font(bold=True, size=13, color=C_FG, name="Arial")
tc.fill      = PatternFill("solid", start_color=C_HDR)
tc.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 26

for col, h in enumerate(
    ["#","Date","Day","Day Type","Time","Predicted Load (MW)","Period"], 1
):
    hdr(ws1.cell(row=2, column=col, value=h), bg=C_SUB)

for i, (ts, pred) in enumerate(zip(future_times, all_predictions), 1):
    bg    = row_bg(ts)
    dtype = "Weekend" if is_weekend(ts) else "Weekday"
    for col, val in enumerate([
        i, ts.strftime("%Y-%m-%d"), ts.strftime("%A"), dtype,
        ts.strftime("%H:%M"), round(float(pred), 2), get_period_name(ts.hour),
    ], 1):
        dat(ws1.cell(row=i + 2, column=col, value=val), bg=bg)

sr = total_slots + 4
ws1.merge_cells(f"A{sr}:G{sr}")
sc = ws1.cell(row=sr, column=1, value="DAILY SUMMARY")
sc.font = Font(bold=True, size=11, color=C_FG, name="Arial")
sc.fill = PatternFill("solid", start_color=C_HDR)
sc.alignment = Alignment(horizontal="center", vertical="center")
for col, h in enumerate(
    ["Day","Date","Day Type","Min (MW)","Max (MW)","Avg (MW)","Total (MWh)"], 1
):
    hdr(ws1.cell(row=sr+1, column=col, value=h), bg=C_SUB)
for d in range(PREDICT_DAYS):
    dp  = all_predictions[d*96:(d+1)*96]
    dd  = start_date + timedelta(days=d)
    clr = DAY_COLORS[d % len(DAY_COLORS)]
    dtype = "Weekend" if is_weekend(dd) else "Weekday"
    for col, val in enumerate([
        f"Day {d+1}", dd.strftime("%Y-%m-%d (%A)"), dtype,
        round(float(dp.min()),2), round(float(dp.max()),2),
        round(float(dp.mean()),2), round(float(dp.sum()*0.25),2),
    ], 1):
        c = ws1.cell(row=sr+2+d, column=col, value=val)
        c.font = Font(bold=(col==1), size=10, name="Arial",
                      color=C_FG if col==1 else "000000")
        c.fill = PatternFill("solid", start_color=clr if col==1 else "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
for i, w in enumerate([6,14,14,12,10,22,12], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for d in range(PREDICT_DAYS):
    dd     = start_date + timedelta(days=d)
    dp     = all_predictions[d*96:(d+1)*96]
    dt_arr = future_times[d*96:(d+1)*96]
    clr    = DAY_COLORS[d % len(DAY_COLORS)]
    dtype  = "Weekend" if is_weekend(dd) else "Weekday"

    ws = wb.create_sheet(dd.strftime("%d %b"))
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = (f"{dd.strftime('%A, %d %B %Y')}  ·  Day {d+1}  ·  {dtype}"
                   f"  ·  {calendar.month_name[dd.month]} model  ·  Optuna-tuned")
    t.font      = Font(bold=True, size=12, color=C_FG, name="Arial")
    t.fill      = PatternFill("solid", start_color=clr)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(
        ["#","Time","Period","Day Type","Predicted Load (MW)","vs Day Avg (MW)"], 1
    ):
        hdr(ws.cell(row=2, column=col, value=h), bg=C_SUB)

    day_avg = float(dp.mean())
    for i, (ts, pred) in enumerate(zip(dt_arr, dp), 1):
        bg = row_bg(ts)
        for col, val in enumerate([
            i, ts.strftime("%H:%M"), get_period_name(ts.hour),
            "Weekend" if is_weekend(ts) else "Weekday",
            round(float(pred), 2), round(float(pred) - day_avg, 2),
        ], 1):
            dat(ws.cell(row=i + 2, column=col, value=val), bg=bg)

    for col, (lbl, val) in enumerate(zip(
        ["Min (MW)","Max (MW)","Avg (MW)","Total (MWh)","Peak Time","Day Type"],
        [round(float(dp.min()),2), round(float(dp.max()),2),
         round(float(dp.mean()),2), round(float(dp.sum()*0.25),2),
         dt_arr[int(np.argmax(dp))].strftime("%H:%M"), dtype]
    ), 1):
        hdr(ws.cell(row=99, column=col, value=lbl), bg=clr)
        dat(ws.cell(row=100, column=col, value=val))

    for i, w in enumerate([5,10,14,12,22,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    chart = LineChart()
    chart.title        = f"{dd.strftime('%d %b %Y')} — {dtype}"
    chart.y_axis.title = "Load (MW)";  chart.x_axis.title = "Time"
    chart.style = 10;  chart.width = 24;  chart.height = 14
    data_ref  = Reference(ws, min_col=5, min_row=2, max_row=98)
    label_ref = Reference(ws, min_col=2, min_row=3, max_row=98)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(label_ref)
    chart.series[0].graphicalProperties.line.solidFill = clr
    chart.series[0].graphicalProperties.line.width     = 18000
    ws.add_chart(chart, "H2")
ws_m = wb.create_sheet("Model Performance")
ws_m.merge_cells("A1:H1")
tm = ws_m["A1"]
tm.value     = f"Optuna-Tuned Context Models — Best Parameters & Accuracy  ({N_TRIALS} trials each)"
tm.font      = Font(bold=True, size=13, color=C_FG, name="Arial")
tm.fill      = PatternFill("solid", start_color=C_HDR)
tm.alignment = Alignment(horizontal="center", vertical="center")
ws_m.row_dimensions[1].height = 26

for col, h in enumerate(
    ["Month","Period","Day Type","Training Rows",
     "CV RMSE (MW)","Val MAE (MW)","Val MAPE (%)","Best n_estimators"], 1
):
    hdr(ws_m.cell(row=2, column=col, value=h), bg=C_SUB)

for i, ((month, period, day_type), meta) in enumerate(sorted(model_meta.items()), 3):
    bg = "DEEAF1" if i % 2 == 0 else "FFFFFF"
    for col, val in enumerate([
        calendar.month_name[month], period, day_type.title(),
        meta["rows"], meta["cv_rmse"], meta["val_mae"], meta["val_mape"],
        meta["params"].get("n_estimators", "—"),
    ], 1):
        dat(ws_m.cell(row=i, column=col, value=val), bg=bg)
detail_start = len(model_meta) + 5
ws_m.merge_cells(f"A{detail_start}:H{detail_start}")
dp_cell = ws_m.cell(row=detail_start, column=1, value="BEST HYPERPARAMETERS PER CONTEXT")
dp_cell.font = Font(bold=True, size=11, color=C_FG, name="Arial")
dp_cell.fill = PatternFill("solid", start_color=C_HDR)
dp_cell.alignment = Alignment(horizontal="center", vertical="center")

param_keys = ["n_estimators","learning_rate","max_depth","min_child_weight",
              "subsample","colsample_bytree","reg_alpha","reg_lambda","gamma"]
for col, h in enumerate(["Context"] + param_keys, 1):
    hdr(ws_m.cell(row=detail_start+1, column=col, value=h), bg=C_SUB)

for i, ((month, period, day_type), meta) in enumerate(sorted(model_meta.items()),
                                                        detail_start + 2):
    bg  = "DEEAF1" if i % 2 == 0 else "FFFFFF"
    ctx = f"{calendar.month_abbr[month]} / {period} / {day_type.title()}"
    dat(ws_m.cell(row=i, column=1, value=ctx), bg=bg)
    for col, pk in enumerate(param_keys, 2):
        v = meta["params"].get(pk, "—")
        dat(ws_m.cell(row=i, column=col,
                      value=round(v, 5) if isinstance(v, float) else v), bg=bg)

for i, w in enumerate([28,12,12,16,14,14,14,16]+[14]*len(param_keys), 1):
    ws_m.column_dimensions[get_column_letter(i)].width = w
ws_t = wb.create_sheet("Tuning History")
ws_t.merge_cells("A1:K1")
th = ws_t["A1"]
th.value     = f"Optuna Trial History — All {N_TRIALS} Trials per Context"
th.font      = Font(bold=True, size=13, color=C_FG, name="Arial")
th.fill      = PatternFill("solid", start_color=C_HDR)
th.alignment = Alignment(horizontal="center", vertical="center")
ws_t.row_dimensions[1].height = 26

trial_cols = ["Context","Trial","CV RMSE","n_estimators","learning_rate",
              "max_depth","min_child_weight","subsample",
              "colsample_bytree","reg_alpha","reg_lambda"]
for col, h in enumerate(trial_cols, 1):
    hdr(ws_t.cell(row=2, column=col, value=h), bg=C_SUB)

row_idx = 3
for (month, period, day_type), trials in sorted(all_trials.items()):
    ctx_label = f"{calendar.month_abbr[month]}/{period}/{day_type.title()}"
    best_rmse = min(t["cv_rmse"] for t in trials)
    for t in trials:
        bg = "E2EFDA" if t["cv_rmse"] == best_rmse else (
             "DEEAF1" if row_idx % 2 == 0 else "FFFFFF")
        for col, val in enumerate([
            ctx_label, t["trial"], t["cv_rmse"],
            t.get("n_estimators",""), t.get("learning_rate",""),
            t.get("max_depth",""), t.get("min_child_weight",""),
            t.get("subsample",""), t.get("colsample_bytree",""),
            t.get("reg_alpha",""), t.get("reg_lambda",""),
        ], 1):
            dat(ws_t.cell(row=row_idx, column=col, value=val), bg=bg)
        row_idx += 1

for i, w in enumerate([28,8,12,14,16,12,18,12,16,12,12], 1):
    ws_t.column_dimensions[get_column_letter(i)].width = w

wb.save(output_file)
print(f"\n[6/6] Done!  Saved → {output_file}")
print(f"\n{'='*60}")
print(f"  Total energy ({PREDICT_DAYS}d) : {all_predictions.sum() * 0.25:.1f} MWh")
print(f"  Overall peak          : {all_predictions.max():.1f} MW  "
      f"@ {future_times[int(np.argmax(all_predictions))].strftime('%Y-%m-%d %H:%M')}")
print(f"  Overall min           : {all_predictions.min():.1f} MW")
print(f"{'='*60}")